from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# Configuration
# =============================================================================

DEFAULT_BENCHMARK_ROOT = Path("benchmarks")
DEFAULT_WORKBOOK_PATH = Path(r"..\..\..\Conference\Metrics.xlsx")
PRUNED_BENCHMARK_ROOT = Path("benchmarks/pruned_models")
PRUNED_WORKBOOK_PATH = Path(r"..\..\..\Conference\Pruning_Metrics.xlsx")
SHEET_NAME = "Metrics Tracker"
HEADER_ROW = 4

CLASSIFIER_MAP = {
    "motionbert": "MotionBERT",
    "cnnlstm": "CNN-LSTM",
    "stgcn": "ST-GCN",
}

ARCHITECTURE_MAP = {
    "alphapose": "Alphapose",
    "alpha_pose": "Alphapose",
    "vitpose": "ViTPose",
    "vit_pose": "ViTPose",
}

VERSION_MAP = {
    "base": "Base",
    "fp32": "FP32",
    "fp16": "FP16",
    "int8": "INT8",
    "original": "Base",
    "fp32_fp32": "FP32",
    "fp16det_fp32pose": "FP32",
    "fp16_fp16": "FP16",
    "int8det_fp16pose": "INT8",
}

PRUNED_POSE_MODELS = {
    "yolo11n",
    "yolo11s",
    "yolo11m",
    "yolo11l",
    "yolo11x",
}

PRUNED_VERSION_FOLDERS = {
    "pruned_80",
    "pruned_90",
}

PRUNED_WRITABLE_METRIC_KEYS = {
    "fps",
    "inference_time",
    "cpu_usage",
    "gpu_usage",
    "memory_usage",
    "temperature",
    "power_draw",
}


# =============================================================================
# Data classes
# =============================================================================

@dataclass
class BenchmarkRecord:
    pose_model_raw: str
    pose_model_excel: str
    version_folder: str
    version_excel: str
    classifier_raw: str
    classifier_excel: str
    run_dir: Path
    summary_json: Path


@dataclass
class Metrics:
    avg_fps: Optional[float]
    inference_time_per_frame_ms: Optional[float]
    avg_cpu_pct: Optional[float]
    avg_gpu_pct: Optional[float]
    avg_ram_pct: Optional[float]
    avg_temp_c: Optional[float]
    avg_power_w: Optional[float]


@dataclass
class Stats:
    folders_scanned: int = 0
    rows_updated: int = 0
    missing_summary: int = 0
    parse_failures: int = 0
    no_matching_row: int = 0
    json_failures: int = 0
    workbook_write_failures: int = 0


@dataclass(frozen=True)
class ModeConfig:
    name: str
    benchmark_root: Path
    workbook_path: Path
    writable_metric_keys: Optional[Set[str]] = None


# =============================================================================
# Logging helpers
# =============================================================================

def info(message: str) -> None:
    print(f"INFO: {message}")


def warn(message: str) -> None:
    print(f"WARNING: {message}")


# =============================================================================
# Normalization helpers
# =============================================================================

def normalize_text(value: object) -> str:
    """
    Normalize workbook/data values for matching:
    - trim
    - lowercase
    - remove spaces, hyphens, underscores
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s\-_]+", "", text)
    return text


def normalize_header(value: object) -> str:
    """
    Normalize headers more aggressively than row values.
    This helps tolerate:
    - extra spaces
    - line breaks
    - punctuation
    - symbols like %, (), /
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


# Canonical internal keys for headers
HEADER_ALIASES = {
    # identifying columns
    "posemodel": "pose_model",
    "version": "version",
    "classifier": "classifier",

    # metric columns
    "fps": "fps",
    "inferencetimeframems": "inference_time",
    "inferencetimeperframems": "inference_time",
    "inferencetimeframe": "inference_time",
    "cpuusage": "cpu_usage",
    "cpuusagepct": "cpu_usage",
    "gpuusage": "gpu_usage",
    "gpuusagepct": "gpu_usage",
    "memoryusage": "memory_usage",
    "memoryusagepct": "memory_usage",
    "memoryusagepercent": "memory_usage",
    "temperaturec": "temperature",
    "temperature": "temperature",
    "powerdraww": "power_draw",
    "powerdraw": "power_draw",
}


# =============================================================================
# JSON helpers
# =============================================================================

def get_nested_number(data: dict, *keys: str) -> Optional[float]:
    current = data
    for key in keys:
        if not isinstance(current, dict) or key not in current:
            return None
        current = current[key]

    try:
        if current is None:
            return None
        return float(current)
    except (TypeError, ValueError):
        return None


def compute_temporal_ms_per_frame(summary: dict) -> Optional[float]:
    """
    Best-effort temporal classifier cost per raw frame.

    Priority:
    1) avg_temporal_effective_ms_per_frame (already amortized)
    2) avg_temporal_total_ms_per_window * num_windows_evaluated / num_frames_processed
    """
    temporal_effective = get_nested_number(summary, "avg_temporal_effective_ms_per_frame")
    if temporal_effective is not None and temporal_effective > 0.0:
        return temporal_effective

    temporal_per_window = get_nested_number(summary, "avg_temporal_total_ms_per_window")
    windows = get_nested_number(summary, "num_windows_evaluated")
    frames = get_nested_number(summary, "num_frames_processed")
    if (
        temporal_per_window is not None
        and windows is not None
        and frames is not None
        and frames > 0.0
    ):
        return (temporal_per_window * windows) / frames

    return None


def compute_inference_time_per_frame_ms(summary: dict) -> Optional[float]:
    """
    Compute the workbook's "Inference Time/Frame (ms)" robustly across
    legacy and new benchmark summary formats.
    """
    # New shared wrapper summaries expose this directly.
    total_loop_ms = get_nested_number(summary, "avg_total_loop_ms_per_frame")
    if total_loop_ms is not None and total_loop_ms > 0.0:
        return total_loop_ms

    # Fallback for new-format components when total loop is absent.
    pose_ms = get_nested_number(summary, "avg_pose_infer_ms_per_frame")
    track_ms = get_nested_number(summary, "avg_track_ms_per_frame")
    render_ms = get_nested_number(summary, "avg_render_ms_per_frame")
    writer_ms = get_nested_number(summary, "avg_writer_ms_per_frame")
    temporal_ms = compute_temporal_ms_per_frame(summary)
    if None not in (pose_ms, track_ms, render_ms, writer_ms, temporal_ms):
        return pose_ms + track_ms + render_ms + writer_ms + temporal_ms

    # Legacy format compatibility.
    preprocess = get_nested_number(summary, "preprocess_ms", "mean")
    inference = get_nested_number(summary, "inference_ms", "mean")
    postprocess = get_nested_number(summary, "postprocess_ms", "mean")
    if preprocess is not None and inference is not None and postprocess is not None:
        return preprocess + inference + postprocess

    # Last resort.
    avg_fps = get_nested_number(summary, "avg_fps")
    if avg_fps is not None and avg_fps > 0.0:
        return 1000.0 / avg_fps

    return None


def extract_metrics(summary: dict) -> Metrics:
    inference_time_per_frame_ms = compute_inference_time_per_frame_ms(summary)

    cpu_temp = get_nested_number(summary, "avg_cpu_temp_c")
    gpu_temp = get_nested_number(summary, "avg_gpu_temp_c")
    avg_temp_c = None
    if cpu_temp is not None and gpu_temp is not None:
        avg_temp_c = (cpu_temp + gpu_temp) / 2.0

    return Metrics(
        avg_fps=get_nested_number(summary, "avg_fps"),
        inference_time_per_frame_ms=inference_time_per_frame_ms,
        avg_cpu_pct=get_nested_number(summary, "avg_cpu_pct"),
        avg_gpu_pct=get_nested_number(summary, "avg_gpu_pct"),
        avg_ram_pct=get_nested_number(summary, "avg_ram_pct"),
        avg_temp_c=avg_temp_c,
        avg_power_w=get_nested_number(summary, "avg_power_w"),
    )


# =============================================================================
# Folder parsing
# =============================================================================

def parse_classifier(run_dir_name: str) -> Optional[str]:
    lower_name = run_dir_name.lower()

    match = re.search(r"__model_(motionbert|cnnlstm|stgcn)\b", lower_name)
    if match:
        return match.group(1)

    for key in CLASSIFIER_MAP:
        if key in lower_name:
            return key

    return None


def map_regular_pose_model(folder_name: str) -> str:
    normalized = folder_name.strip()
    lowered = normalized.lower()
    return ARCHITECTURE_MAP.get(lowered, normalized)


def map_pose_model(folder_name: str, mode: str) -> Optional[str]:
    if mode == "regular":
        return map_regular_pose_model(folder_name)

    if mode == "pruned":
        lowered = folder_name.strip().lower()
        if lowered not in PRUNED_POSE_MODELS:
            return None
        return f"{lowered}-pose"

    raise ValueError(f"Unsupported mode: {mode}")


def map_version(version_folder_name: str, mode: str) -> Optional[str]:
    if mode == "regular":
        return VERSION_MAP.get(version_folder_name.strip().lower())

    if mode == "pruned":
        lowered = version_folder_name.strip().lower()
        if lowered in PRUNED_VERSION_FOLDERS:
            return lowered
        return None

    raise ValueError(f"Unsupported mode: {mode}")


def scan_benchmark_records(
    benchmark_root: Path,
    stats: Stats,
    mode: str,
) -> List[BenchmarkRecord]:
    records: List[BenchmarkRecord] = []

    if not benchmark_root.exists():
        raise FileNotFoundError(f"Benchmark root not found: {benchmark_root}")
    if not benchmark_root.is_dir():
        raise NotADirectoryError(f"Benchmark root is not a directory: {benchmark_root}")

    for pose_model_dir in sorted(p for p in benchmark_root.iterdir() if p.is_dir()):
        pose_model_raw = pose_model_dir.name
        pose_model_excel = map_pose_model(pose_model_raw, mode)
        if pose_model_excel is None:
            warn(f"Skipping unknown pose model folder for mode={mode}: {pose_model_dir}")
            stats.parse_failures += 1
            continue

        for version_dir in sorted(p for p in pose_model_dir.iterdir() if p.is_dir()):
            version_folder = version_dir.name
            version_excel = map_version(version_folder, mode)
            if version_excel is None:
                warn(f"Skipping unknown version folder for mode={mode}: {version_dir}")
                stats.parse_failures += 1
                continue

            for run_dir in sorted(p for p in version_dir.iterdir() if p.is_dir()):
                stats.folders_scanned += 1

                summary_json = run_dir / "summary.json"
                if not summary_json.exists():
                    warn(f"Missing summary.json: {summary_json}")
                    stats.missing_summary += 1
                    continue

                classifier_raw = parse_classifier(run_dir.name)
                if classifier_raw is None:
                    warn(f"Could not parse classifier from folder name: {run_dir.name}")
                    stats.parse_failures += 1
                    continue

                classifier_excel = CLASSIFIER_MAP.get(classifier_raw)
                if classifier_excel is None:
                    warn(f"Classifier parsed but not mapped: {classifier_raw!r}")
                    stats.parse_failures += 1
                    continue

                records.append(
                    BenchmarkRecord(
                        pose_model_raw=pose_model_raw,
                        pose_model_excel=pose_model_excel,
                        version_folder=version_folder,
                        version_excel=version_excel,
                        classifier_raw=classifier_raw,
                        classifier_excel=classifier_excel,
                        run_dir=run_dir,
                        summary_json=summary_json,
                    )
                )

    return records


# =============================================================================
# Workbook helpers
# =============================================================================

def resolve_header_columns(ws: Worksheet) -> Dict[str, int]:
    """
    Read row HEADER_ROW and map workbook headers to canonical internal keys.

    This is more robust than exact string matching and will tolerate
    minor header text differences like:
    - "Memory Usage %"
    - "Memory Usage (%)"
    - line breaks in header text
    """
    resolved: Dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=HEADER_ROW, column=col).value
        norm = normalize_header(raw)
        canonical = HEADER_ALIASES.get(norm)
        if canonical and canonical not in resolved:
            resolved[canonical] = col

    required = ["pose_model", "version", "classifier"]
    missing = [key for key in required if key not in resolved]
    if missing:
        raise KeyError(
            f"Missing required headers on row {HEADER_ROW}: {missing}. "
            f"Check the sheet header text."
        )

    return resolved


def build_row_index(ws: Worksheet, columns: Dict[str, int]) -> Dict[Tuple[str, str, str], int]:
    """
    Build an index of workbook rows.

    Important:
    - carries forward Pose Model / Version / Classifier values when cells are blank
    - this handles grouped/merged Excel layouts where repeated identifiers only
      appear once visually
    """
    pose_col = columns["pose_model"]
    version_col = columns["version"]
    classifier_col = columns["classifier"]

    row_index: Dict[Tuple[str, str, str], int] = {}

    current_pose = ""
    current_version = ""
    current_classifier = ""

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        pose_raw = ws.cell(row=row, column=pose_col).value
        version_raw = ws.cell(row=row, column=version_col).value
        classifier_raw = ws.cell(row=row, column=classifier_col).value

        # Update carry-forward values only when a non-empty cell appears
        if pose_raw not in (None, ""):
            current_pose = str(pose_raw).strip()
        if version_raw not in (None, ""):
            current_version = str(version_raw).strip()
        if classifier_raw not in (None, ""):
            current_classifier = str(classifier_raw).strip()

        # If the classifier cell is blank on this actual row, it usually means
        # this is not a real data row for a unique combination, so skip it.
        if classifier_raw in (None, ""):
            continue

        key = (
            normalize_text(current_pose),
            normalize_text(current_version),
            normalize_text(current_classifier),
        )

        if not all(key):
            continue

        if key in row_index:
            warn(
                f"Duplicate workbook key at row {row} for "
                f"{current_pose} | {current_version} | {current_classifier}. "
                f"Using the later row."
            )

        row_index[key] = row

    return row_index


def find_matching_row(
    row_index: Dict[Tuple[str, str, str], int],
    pose_model: str,
    version_excel: str,
    classifier_excel: str,
) -> Optional[int]:
    key = (
        normalize_text(pose_model),
        normalize_text(version_excel),
        normalize_text(classifier_excel),
    )
    return row_index.get(key)


def write_metrics_to_row(
    ws: Worksheet,
    columns: Dict[str, int],
    row_number: int,
    metrics: Metrics,
    writable_metric_keys: Optional[Set[str]] = None,
) -> None:
    metric_values = {
        "fps": metrics.avg_fps,
        "inference_time": metrics.inference_time_per_frame_ms,
        "cpu_usage": metrics.avg_cpu_pct,
        "gpu_usage": metrics.avg_gpu_pct,
        "memory_usage": metrics.avg_ram_pct,
        "temperature": metrics.avg_temp_c,
        "power_draw": metrics.avg_power_w,
    }

    for key, value in metric_values.items():
        if writable_metric_keys is not None and key not in writable_metric_keys:
            continue
        col = columns.get(key)
        if col is None:
            warn(f"Metric column not found in workbook header row for key: {key}")
            continue
        if value is None:
            continue
        ws.cell(row=row_number, column=col).value = value


# =============================================================================
# Main update logic
# =============================================================================

def update_workbook(mode_config: ModeConfig) -> Stats:
    stats = Stats()
    records = scan_benchmark_records(mode_config.benchmark_root, stats, mode_config.name)
    info(f"Found {len(records)} benchmark folders eligible for processing.")

    workbook_path = mode_config.workbook_path
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    columns = resolve_header_columns(ws)
    row_index = build_row_index(ws, columns)

    info(f"Indexed {len(row_index)} workbook data rows.")

    for record in records:
        try:
            with record.summary_json.open("r", encoding="utf-8") as f:
                summary = json.load(f)
        except Exception as exc:
            warn(f"Failed to read JSON {record.summary_json}: {exc}")
            stats.json_failures += 1
            continue

        row_number = find_matching_row(
            row_index=row_index,
            pose_model=record.pose_model_excel,
            version_excel=record.version_excel,
            classifier_excel=record.classifier_excel,
        )

        if row_number is None:
            warn(
                "No matching Excel row for "
                f"pose_model={record.pose_model_excel!r}, "
                f"pose_model_raw={record.pose_model_raw!r}, "
                f"version={record.version_excel!r}, "
                f"version_folder={record.version_folder!r}, "
                f"classifier={record.classifier_excel!r} "
                f"(folder: {record.run_dir})"
            )
            stats.no_matching_row += 1
            continue

        try:
            metrics = extract_metrics(summary)
            write_metrics_to_row(
                ws,
                columns,
                row_number,
                metrics,
                writable_metric_keys=mode_config.writable_metric_keys,
            )
            stats.rows_updated += 1
            info(
                f"Updated row {row_number}: "
                f"{record.pose_model_excel} | {record.version_excel} | {record.classifier_excel}"
            )
        except Exception as exc:
            warn(f"Failed to update workbook for {record.run_dir}: {exc}")
            stats.workbook_write_failures += 1
            continue

    try:
        wb.save(workbook_path)
    except PermissionError:
        raise PermissionError(
            f"Cannot save workbook because it is locked: {workbook_path}\n"
            f"Close the Excel file if it is open, close Explorer preview, and if needed "
            f"pause OneDrive sync briefly, then run again."
        )

    return stats


def print_summary(stats: Stats) -> None:
    print("\n" + "=" * 60)
    print("Update Summary")
    print("=" * 60)
    print(f"Total benchmark folders scanned:      {stats.folders_scanned}")
    print(f"Successfully updated rows:           {stats.rows_updated}")
    print(f"Skipped (missing summary.json):      {stats.missing_summary}")
    print(f"Skipped (parse failures):            {stats.parse_failures}")
    print(f"Skipped (JSON read failures):        {stats.json_failures}")
    print(f"Skipped (no matching Excel row):     {stats.no_matching_row}")
    print(f"Skipped (workbook write failures):   {stats.workbook_write_failures}")


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Write benchmark metrics from summary.json files into the Excel tracker workbook."
    )
    parser.add_argument(
        "--mode",
        choices=("regular", "pruned"),
        default="regular",
        help="Workbook/update mode. 'regular' preserves the current default behavior.",
    )
    return parser.parse_args(argv)


def resolve_mode_config(mode: str) -> ModeConfig:
    if mode == "regular":
        return ModeConfig(
            name="regular",
            benchmark_root=DEFAULT_BENCHMARK_ROOT.resolve(),
            workbook_path=DEFAULT_WORKBOOK_PATH.resolve(),
        )

    if mode == "pruned":
        return ModeConfig(
            name="pruned",
            benchmark_root=PRUNED_BENCHMARK_ROOT.resolve(),
            workbook_path=PRUNED_WORKBOOK_PATH.resolve(),
            writable_metric_keys=PRUNED_WRITABLE_METRIC_KEYS,
        )

    raise ValueError(f"Unsupported mode: {mode}")


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    mode_config = resolve_mode_config(args.mode)

    info(f"Mode:           {mode_config.name}")
    info(f"Benchmark root: {mode_config.benchmark_root}")
    info(f"Workbook path:  {mode_config.workbook_path}")
    info(f"Sheet name:     {SHEET_NAME}")

    try:
        stats = update_workbook(mode_config)
        print_summary(stats)
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
